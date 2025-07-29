
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
import sqlite3
from datetime import datetime
import atexit
from openpyxl import Workbook, load_workbook
import os
import smtplib
from email.message import EmailMessage

# === Configuración de correo ===
EMAIL_EMISOR = "sistemamantenimientoalertas@gmail.com"
CLAVE_APP = "slucdkpwoxibxfym"

def enviar_alerta_por_correo(texto_alerta, lista_destinatarios):
    msg = EmailMessage()
    msg['Subject'] = "🔧 Alertas de Mantenimiento de Tornos"
    msg['From'] = EMAIL_EMISOR
    msg['To'] = ", ".join(lista_destinatarios)
    msg.set_content(texto_alerta)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_EMISOR, CLAVE_APP)
            smtp.send_message(msg)
        print("✅ Correo enviado correctamente.")
    except Exception as e:
        print(f"❌ Error al enviar el correo: {e}")

# === Configuración general del sistema ===
excel_file = "mantenimientos.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"
    ws.append(["Torno ID", "Fecha", "Tipo", "Descripción", "Repuestos", "Técnico"])
    wb.save(excel_file)

def obtener_hoja_mantenimientos():
    wb = load_workbook(excel_file)
    hoja_objetivo = "Mantenimientos"
    if hoja_objetivo not in wb.sheetnames:
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            ws.title = hoja_objetivo
        else:
            ws = wb.create_sheet(title=hoja_objetivo)
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            ws.append(["Torno ID", "Fecha", "Tipo", "Descripción", "Repuestos", "Técnico"])
        wb.save(excel_file)
    return load_workbook(excel_file)[hoja_objetivo]

conn = sqlite3.connect('mantenimiento_torno.db')
cursor = conn.cursor()

def cerrar_conexion():
    conn.close()

atexit.register(cerrar_conexion)

cursor.execute("""CREATE TABLE IF NOT EXISTS mantenimiento (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    torno_id TEXT,
    fecha TEXT,
    tipo TEXT,
    descripcion TEXT,
    repuestos TEXT,
    tecnico TEXT
)""")
cursor.execute("""CREATE TABLE IF NOT EXISTS programacion (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    torno_id TEXT,
    fecha_proxima TEXT,
    tipo TEXT
)""")
conn.commit()

def sincronizar_excel_a_sqlite():
    ws = obtener_hoja_mantenimientos()
    filas = list(ws.iter_rows(min_row=2, values_only=True))

    for fila in filas:
        if not fila or all(f is None for f in fila):
            continue
        fila = list(fila)
        if len(fila) < 6:
            print(f"[Advertencia] Fila incompleta detectada y corregida: {fila}")
            while len(fila) < 6:
                fila.append("N/A")

        torno_id, fecha, tipo, descripcion, repuestos, tecnico = fila

        cursor.execute("""
            SELECT * FROM mantenimiento 
            WHERE torno_id = ? AND fecha = ? AND tipo = ? AND descripcion = ? AND repuestos = ? AND tecnico = ?
        """, (torno_id, fecha, tipo, descripcion, repuestos, tecnico))

        existe = cursor.fetchone()
        if not existe:
            cursor.execute("""
                INSERT INTO mantenimiento (torno_id, fecha, tipo, descripcion, repuestos, tecnico)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (torno_id, fecha, tipo, descripcion, repuestos, tecnico))

    conn.commit()

def validar_fecha(fecha_texto):
    try:
        datetime.strptime(fecha_texto, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def registrar_mantenimiento():
    def guardar():
        if not validar_fecha(entry_fecha.get()):
            messagebox.showerror("Error", "Fecha inválida. Use formato YYYY-MM-DD.")
            return
        datos = (
            entry_torno.get(),
            entry_fecha.get(),
            combo_tipo.get().lower(),
            entry_desc.get(),
            entry_repuestos.get(),
            entry_tecnico.get()
        )
        cursor.execute("""INSERT INTO mantenimiento 
                          (torno_id, fecha, tipo, descripcion, repuestos, tecnico) 
                          VALUES (?, ?, ?, ?, ?, ?)""", datos)
        conn.commit()
        ws = obtener_hoja_mantenimientos()
        fila = list(datos)
        ws.append(fila)
        wb = ws.parent
        wb.save(excel_file)
        messagebox.showinfo("Éxito", "Mantenimiento registrado.")
        ventana.destroy()

    ventana = tk.Toplevel(root)
    ventana.title("Registrar Mantenimiento")
    tk.Label(ventana, text="Torno ID").grid(row=0, column=0)
    tk.Label(ventana, text="Fecha (YYYY-MM-DD)").grid(row=1, column=0)
    tk.Label(ventana, text="Tipo").grid(row=2, column=0)
    tk.Label(ventana, text="Descripción").grid(row=3, column=0)
    tk.Label(ventana, text="Repuestos").grid(row=4, column=0)
    tk.Label(ventana, text="Técnico").grid(row=5, column=0)

    entry_torno = tk.Entry(ventana)
    entry_fecha = tk.Entry(ventana)
    entry_fecha.insert(0, datetime.today().strftime('%Y-%m-%d'))
    combo_tipo = ttk.Combobox(ventana, values=["preventivo", "correctivo"])
    combo_tipo.set("preventivo")
    entry_desc = tk.Entry(ventana)
    entry_repuestos = tk.Entry(ventana)
    entry_tecnico = tk.Entry(ventana)

    entry_torno.grid(row=0, column=1)
    entry_fecha.grid(row=1, column=1)
    combo_tipo.grid(row=2, column=1)
    entry_desc.grid(row=3, column=1)
    entry_repuestos.grid(row=4, column=1)
    entry_tecnico.grid(row=5, column=1)

    tk.Button(ventana, text="Guardar", command=guardar).grid(row=6, column=0, columnspan=2)

def consultar_historial():
    opcion = simpledialog.askstring("Consulta", "Consultar por:\n1. ID del torno\n2. Rango de fechas (1/2):")
    if opcion == '1':
        torno_id = simpledialog.askstring("Consulta", "Ingrese ID del torno:")
        cursor.execute('SELECT * FROM mantenimiento WHERE torno_id = ?', (torno_id,))
    elif opcion == '2':
        fecha_inicio = simpledialog.askstring("Consulta", "Fecha inicio (YYYY-MM-DD):")
        fecha_fin = simpledialog.askstring("Consulta", "Fecha fin (YYYY-MM-DD):")
        if not (validar_fecha(fecha_inicio) and validar_fecha(fecha_fin)):
            messagebox.showerror("Error", "Fechas inválidas.")
            return
        cursor.execute('SELECT * FROM mantenimiento WHERE fecha BETWEEN ? AND ?', (fecha_inicio, fecha_fin))
    else:
        messagebox.showerror("Error", "Opción inválida.")
        return

    registros = cursor.fetchall()
    if registros:
        resultado = ""
        for r in registros:
            resultado += f"ID:{r[0]} Torno:{r[1]} Fecha:{r[2]} Tipo:{r[3]} Descripción:{r[4]} Repuestos:{r[5]} Técnico:{r[6]}\n"
        mostrar_resultado("Historial de Mantenimiento", resultado)
    else:
        messagebox.showinfo("Resultado", "No se encontraron registros.")

def programar_mantenimiento():
    def guardar():
        if not validar_fecha(entry_fecha.get()):
            messagebox.showerror("Error", "Fecha inválida. Use formato YYYY-MM-DD.")
            return
        datos = (
            entry_torno.get(),
            entry_fecha.get(),
            combo_tipo.get().lower()
        )
        cursor.execute('INSERT INTO programacion (torno_id, fecha_proxima, tipo) VALUES (?, ?, ?)', datos)
        conn.commit()
        messagebox.showinfo("Éxito", "Mantenimiento programado.")
        ventana.destroy()

    ventana = tk.Toplevel(root)
    ventana.title("Programar Mantenimiento")
    tk.Label(ventana, text="Torno ID").grid(row=0, column=0)
    tk.Label(ventana, text="Fecha Próxima (YYYY-MM-DD)").grid(row=1, column=0)
    tk.Label(ventana, text="Tipo").grid(row=2, column=0)

    entry_torno = tk.Entry(ventana)
    entry_fecha = tk.Entry(ventana)
    entry_fecha.insert(0, datetime.today().strftime('%Y-%m-%d'))
    combo_tipo = ttk.Combobox(ventana, values=["preventivo", "correctivo"])
    combo_tipo.set("preventivo")

    entry_torno.grid(row=0, column=1)
    entry_fecha.grid(row=1, column=1)
    combo_tipo.grid(row=2, column=1)

    tk.Button(ventana, text="Guardar", command=guardar).grid(row=3, column=0, columnspan=2)

def generar_alertas():
    hoy = datetime.today().date()
    alerta_dias = 30
    cursor.execute('SELECT torno_id, fecha_proxima, tipo FROM programacion')
    programaciones = cursor.fetchall()
    alertas = ""
    for torno_id, fecha_str, tipo in programaciones:
        fecha_proxima = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        dias = (fecha_proxima - hoy).days
        if 0 <= dias <= alerta_dias:
            alertas += f"ALERTA: {tipo} para torno {torno_id} el {fecha_str} (en {dias} días)\n"
    if alertas:
        mostrar_resultado("Alertas de Mantenimiento", alertas)
        # Enviar por correo
        correos_destino = ["ismaelpaillacho2006@gmail.com"]
        enviar_alerta_por_correo(alertas, correos_destino)
    else:
        messagebox.showinfo("Alertas", "No hay alertas próximas.")

def mostrar_resultado(titulo, texto):
    ventana = tk.Toplevel(root)
    ventana.title(titulo)
    texto_widget = tk.Text(ventana, width=80, height=20)
    texto_widget.insert(tk.END, texto)
    texto_widget.config(state='disabled')
    texto_widget.pack()

root = tk.Tk()
root.title("Sistema de Mantenimiento de Tornos")
tk.Label(root, text="Seleccione una opción:", font=("Arial", 14)).pack(pady=10)
tk.Button(root, text="Registrar Mantenimiento", width=30, command=registrar_mantenimiento).pack(pady=5)
tk.Button(root, text="Consultar Historial", width=30, command=consultar_historial).pack(pady=5)
tk.Button(root, text="Programar Mantenimiento", width=30, command=programar_mantenimiento).pack(pady=5)
tk.Button(root, text="Generar Alertas", width=30, command=generar_alertas).pack(pady=5)
tk.Button(root, text="Salir", width=30, command=root.destroy).pack(pady=10)

sincronizar_excel_a_sqlite()
root.mainloop()
